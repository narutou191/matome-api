import shutil
from pathlib import Path
from openpyxl import load_workbook
from bot.client_data import ClientData
from bot.config import TEMPLATE_EXCEL, OUTPUT_DIR

class ExcelGenerator:
    def __init__(self):
        self.template_path = TEMPLATE_EXCEL

    def generate(self, client_data: ClientData, filename: str = None) -> Path:
        if not client_data.is_complete():
            raise ValueError("Client data is not complete. All blocks must be filled.")

        if not filename:
            name = client_data.personal.name_full.replace(" ", "_")
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"cliente_{name}_{timestamp}.xlsx"

        output_path = OUTPUT_DIR / filename
        shutil.copy2(self.template_path, output_path)

        wb = load_workbook(output_path)
        ws = wb["Ficha"]

        self._fill_personal_info(ws, client_data.personal)
        self._fill_employment_info(ws, client_data.employment)
        self._fill_family_info(ws, client_data.family)
        self._fill_financing_info(ws, client_data.financing)
        self._fill_special_info(ws, client_data.special)

        wb.save(output_path)
        return output_path

    def _fill_personal_info(self, ws, personal):
        ws["A8"] = personal.name_katakana
        ws["A9"] = personal.name_full
        ws["A12"] = personal.birthdate
        ws["A15"] = personal.address
        ws["A16"] = personal.cep
        ws["A18"] = personal.email
        ws["A19"] = personal.phone
        ws["A21"] = personal.nationality

    def _fill_employment_info(self, ws, employment):
        ws["A47"] = employment.company_name
        ws["A48"] = employment.company_address
        ws["B48"] = employment.company_cep
        ws["A51"] = employment.work_location
        ws["A53"] = employment.work_address
        ws["B53"] = employment.work_cep
        ws["F48"] = employment.annual_income
        ws["F46"] = employment.payment_date
        ws["F55"] = employment.contract_type
        ws["F53"] = employment.hire_date

    def _fill_family_info(self, ws, family):
        ws["E27"] = family.marital_status

        dependent_rows = [
            ("A31", "C31", "D31", "E31", "F31"),
            ("A34", "C34", "D34", "E34", "F34"),
            ("A37", "C37", "D37", "E37", "F37"),
            ("A40", "C40", "D40", "E40", "F40"),
        ]

        for idx, (name_cell, rel_cell, age_cell, income_cell, job_cell) in enumerate(dependent_rows):
            if idx < len(family.dependents):
                dep = family.dependents[idx]
                ws[name_cell] = dep.get("name", "")
                ws[rel_cell] = dep.get("relationship", "")
                ws[age_cell] = dep.get("age", "")
                ws[income_cell] = dep.get("annual_income", "")
                ws[job_cell] = dep.get("job_school", "")

    def _fill_financing_info(self, ws, financing):
        ws["A81"] = "はい" if financing.liquidated_last_3m else "いいえ"
        if financing.liquidated_details:
            ws["C82"] = financing.liquidated_details

        ws["A86"] = "はい" if financing.active_financings else "いいえ"

        start_row = 89
        for idx, fin in enumerate(financing.active_financings[:12]):
            row = start_row + idx
            ws[f"A{row}"] = fin.get("company", "")
            ws[f"B{row}"] = fin.get("purpose", "")
            ws[f"C{row}"] = fin.get("contract_date", "")
            ws[f"D{row}"] = fin.get("amount", "")
            ws[f"E{row}"] = fin.get("monthly_payment", "")
            ws[f"F{row}"] = fin.get("remaining_balance", "")

    def _fill_special_info(self, ws, special):
        ws["A102"] = "はい" if special.has_side_job else "いいえ"
        ws["A107"] = "はい" if special.is_maternity_leave else "いいえ"
        ws["A112"] = "はい" if special.has_existing_illness else "いいえ"
        if special.illness_name:
            ws["C112"] = special.illness_name
        if special.medication_details:
            ws["E112"] = special.medication_details
        if special.additional_notes:
            ws["C42"] = special.additional_notes
